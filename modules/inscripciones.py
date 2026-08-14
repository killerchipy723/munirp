from flask import Blueprint, render_template, request, send_file
import pandas as pd
import io
from db import get_connection

# Importaciones de estilos de openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

inscripciones_bp = Blueprint('inscripciones', __name__)

@inscripciones_bp.route('/admin/inscripciones')
def listar_inscripciones():
    conn = get_connection()
    inscripciones = []
    categorias = []
    circuitos = []
    
    # Capturar filtros de la URL
    filtro_categoria = request.args.get('categoria_id', '')
    filtro_circuito = request.args.get('circuito_id', '')
    filtro_busqueda = request.args.get('busqueda', '').strip()

    try:
        with conn.cursor() as cursor:
            # 1. Obtener listas para los selectores
            cursor.execute("SELECT id, nombre FROM categorias ORDER BY nombre")
            categorias = cursor.fetchall()

            cursor.execute("SELECT id, nombre FROM circuitos ORDER BY nombre")
            circuitos = cursor.fetchall()

            # 2. Armar la consulta principal
            query = """
                SELECT 
                    p.id, p.apellido, p.nombre, p.dni, p.localidad, 
                    p.fecha_nacimiento, p.genero, p.foto,
                    c.nombre AS categoria_nombre, c.grupo,
                    cir.nombre AS circuito_nombre, cir.kilometros
                FROM participantes p
                LEFT JOIN categorias c ON p.categoria_id = c.id
                LEFT JOIN circuitos cir ON p.circuito_id = cir.id
                WHERE 1=1
            """
            params = []

            if filtro_categoria:
                query += " AND p.categoria_id = %s"
                params.append(filtro_categoria)
            
            if filtro_circuito:
                query += " AND p.circuito_id = %s"
                params.append(filtro_circuito)

            if filtro_busqueda:
                query += " AND (p.dni LIKE %s OR p.apellido LIKE %s OR p.nombre LIKE %s OR p.localidad LIKE %s)"
                busq_param = f"%{filtro_busqueda}%"
                params.extend([busq_param, busq_param, busq_param, busq_param])

            query += " ORDER BY p.id DESC"
            
            cursor.execute(query, tuple(params))
            inscripciones = cursor.fetchall()
    finally:
        conn.close()

    return render_template(
        'admin_inscripciones.html',
        inscripciones=inscripciones,
        categorias=categorias,
        circuitos=circuitos,
        filtro_categoria=filtro_categoria,
        filtro_circuito=filtro_circuito,
        filtro_busqueda=filtro_busqueda
    )


# Ruta para exportar a Excel Profesional en A4 Horizontal
@inscripciones_bp.route('/admin/inscripciones/excel')
def exportar_excel():
    conn = get_connection()
    try:
        query = """
            SELECT 
                p.dni AS DNI,
                p.apellido AS Apellido,
                p.nombre AS Nombre,
                p.localidad AS Localidad,
                p.fecha_nacimiento AS 'Fecha de Nacimiento',
                p.genero AS Género,
                c.nombre AS Categoría,
                c.grupo AS Grupo,
                cir.nombre AS Circuito,
                cir.kilometros AS 'Kilómetros'
            FROM participantes p
            LEFT JOIN categorias c ON p.categoria_id = c.id
            LEFT JOIN circuitos cir ON p.circuito_id = cir.id
            ORDER BY p.apellido, p.nombre
        """
        
        with conn.cursor() as cursor:
            cursor.execute(query)
            datos = cursor.fetchall()
            columnas = [desc[0] for desc in cursor.description]
        
        df = pd.DataFrame(datos, columns=columnas)

        # Transformación a mayúsculas
        columnas_a_mayusculas = ['Apellido', 'Nombre', 'Localidad']
        for col in columnas_a_mayusculas:
            if col in df.columns:
                df[col] = df[col].astype(str).str.upper()

        # --- AGREGAR COLUMNA MANUAL DE ASISTENCIA / FIRMA ---
        # Fila vacía con casillero para tildar o firmar manualmente en la acreditación
        df['Asistencia / Firma'] = "[   ]"

    finally:
        conn.close()

    # Generación de Excel en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        start_row = 3
        df.to_excel(writer, index=False, sheet_name='Inscriptos MTB', startrow=start_row)
        
        ws = writer.sheets['Inscriptos MTB']
        ws.views.sheetView[0].showGridLines = True  # Mostrar cuadrícula
        
        # --- CONFIGURACIÓN DE IMPRESIÓN: A4 HORIZONTAL ---
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        
        # Ajustar automáticamente todas las columnas a 1 página de ancho
        ws.page_setup.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.sheet_properties.pageSetUpPr.fitToPage = True

        num_cols = len(df.columns)
        last_col_letter = get_column_letter(num_cols)

        # --- 1. TÍTULO PRINCIPAL ---
        ws.merge_cells(f'A1:{last_col_letter}1')
        title_cell = ws['A1']
        title_cell.value = "2° EDICIÓN DE MOUNTAIN BIKE - RÍO PIEDRAS 2026"
        title_cell.font = Font(name='Calibri', size=15, bold=True, color='FFFFFF')
        title_cell.fill = PatternFill(start_color='1B365D', end_color='1B365D', fill_type='solid') # Azul Marino
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 36

        # --- 2. SUBTÍTULO / FICHA DE CONTROL ---
        ws.merge_cells(f'A2:{last_col_letter}2')
        sub_cell = ws['A2']
        sub_cell.value = "PLANILLA OFICIAL DE ACREDITACIÓN Y CONTROL DE ASISTENCIA"
        sub_cell.font = Font(name='Calibri', size=10, bold=True, italic=True, color='4A5568')
        sub_cell.fill = PatternFill(start_color='EDF2F7', end_color='EDF2F7', fill_type='solid')
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 20

        # Fila 3 vacía como margen
        ws.row_dimensions[3].height = 10

        # --- 3. ESTILOS DE CABECERA DE TABLA (FILA 4) ---
        header_fill = PatternFill(start_color='2B4C7E', end_color='2B4C7E', fill_type='solid')
        header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E0'),
            right=Side(style='thin', color='CBD5E0'),
            top=Side(style='thin', color='CBD5E0'),
            bottom=Side(style='thin', color='CBD5E0')
        )

        header_row_idx = start_row + 1  # Fila 4
        ws.row_dimensions[header_row_idx].height = 28

        for col_num in range(1, num_cols + 1):
            cell = ws.cell(row=header_row_idx, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        # --- 4. ESTILOS DE FILAS DE DATOS ---
        zebra_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid')
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
        data_font = Font(name='Calibri', size=10, color='1A202C')
        
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')

        data_start_row = header_row_idx + 1
        total_rows = len(df)

        for row_idx in range(data_start_row, data_start_row + total_rows):
            ws.row_dimensions[row_idx].height = 24  # Fila un poco más alta para poder firmar con comodidad
            current_fill = zebra_fill if (row_idx % 2 == 0) else white_fill

            for col_num in range(1, num_cols + 1):
                cell = ws.cell(row=row_idx, column=col_num)
                cell.font = data_font
                cell.fill = current_fill
                cell.border = thin_border

                col_name = df.columns[col_num - 1]
                if col_name in ['DNI', 'Fecha de Nacimiento', 'Género', 'Grupo', 'Kilómetros', 'Asistencia / Firma']:
                    cell.alignment = align_center
                else:
                    cell.alignment = align_left

        # --- 5. AJUSTE DE ANCHOS DE COLUMNAS ---
        for col_idx, col_name in enumerate(df.columns, start=1):
            col_letter = get_column_letter(col_idx)
            
            if col_name == 'Asistencia / Firma':
                ws.column_dimensions[col_letter].width = 20  # Ancho generoso para tildar o firmar
            else:
                lengths = [len(str(val)) for val in df[col_name].dropna()] if not df.empty else [0]
                max_val_len = max(lengths) if lengths else 0
                max_len = max(max_val_len, len(str(col_name))) + 4
                ws.column_dimensions[col_letter].width = max(max_len, 11)

    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='Planilla_Acreditacion_MTB_RioPiedras2026.xlsx'
    )